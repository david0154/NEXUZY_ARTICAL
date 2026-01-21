#!/usr/bin/env python3
"""
Export Module - PDF and Excel Export Functionality with Images
Supports exporting articles and user data with embedded images
"""

import os
import sys
from datetime import datetime
from typing import List, Dict
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from PIL import Image
from utils.logger import Logger


class ExportManager:
    """Manages data export to PDF and Excel formats with embedded images"""

    def __init__(self):
        self.logger = Logger(__name__)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Primary color - HEX to RGB conversion for ReportLab
        # #1f77d4 = RGB(31, 119, 212)
        self.PRIMARY_RGB = (31, 119, 212)  # From #1f77d4
        self.PRIMARY_HEX_EXCEL = "1f77d4"
        self.SECONDARY_HEX_EXCEL = "f0f0f0"
        
        # Import image_sync for FTP downloads
        try:
            from utils.image_sync import get_image_sync
            self.image_sync = get_image_sync()
        except Exception as e:
            self.logger.warning(f"Image sync not available: {e}")
            self.image_sync = None

    def _download_image(self, image_path: str, max_width: int = 100, max_height: int = 100) -> str:
        """
        Download image from FTP path and return local cached path
        
        Args:
            image_path: FTP path (e.g., /nexuzy/article_*.jpg)
            max_width: Max width in pixels
            max_height: Max height in pixels
        
        Returns:
            Local file path or None if download fails
        """
        try:
            if not image_path:
                return None
            
            # If it's already a local path, use it directly
            if not image_path.startswith('/'):
                if os.path.exists(image_path):
                    return image_path
                return None
            
            # FTP path - download via image_sync
            if self.image_sync:
                cached_path = self.image_sync.get_cached_path(image_path)
                if cached_path and os.path.exists(cached_path):
                    return cached_path
                
                # Download if not cached
                cached_path = self.image_sync.download_image(image_path)
                if cached_path and os.path.exists(cached_path):
                    return cached_path
            
            return None
        except Exception as e:
            self.logger.debug(f"Image download failed for {image_path}: {e}")
            return None

    def export_articles_to_excel(self, articles: List[Dict], output_path: str = None) -> str:
        """
        Export articles to Excel file with embedded images
        
        Args:
            articles: List of article dictionaries
            output_path: Custom output path (optional)
        
        Returns:
            Path to created Excel file
        """
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Articles"

            # Define styles
            header_fill = PatternFill(start_color=self.PRIMARY_HEX_EXCEL, end_color=self.PRIMARY_HEX_EXCEL, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ["Image", "ID", "Article Name", "Mould", "Size", "Gender", "Created By", "Created Date", "Sync Status"]
            ws.append(headers)

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Set column widths
            ws.column_dimensions['A'].width = 15  # Image column
            ws.column_dimensions['B'].width = 10  # ID
            ws.column_dimensions['C'].width = 20  # Article Name
            ws.column_dimensions['D'].width = 12  # Mould
            ws.column_dimensions['E'].width = 10  # Size
            ws.column_dimensions['F'].width = 12  # Gender
            ws.column_dimensions['G'].width = 12  # Created By
            ws.column_dimensions['H'].width = 15  # Created Date
            ws.column_dimensions['I'].width = 12  # Sync Status

            # Add data rows
            for idx, article in enumerate(articles, start=2):
                # Text data
                row = [
                    "",  # Image placeholder
                    article.get('id', '')[:8],
                    article.get('article_name', ''),
                    article.get('mould', ''),
                    article.get('size', ''),
                    article.get('gender', ''),
                    article.get('created_by', '')[:8],
                    article.get('created_at', '')[:10],
                    "Synced" if article.get('sync_status', 0) == 1 else "Pending"
                ]
                ws.append(row)

                # Format data cells
                for cell in ws[idx]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Try to embed image
                image_path = article.get('image_path')
                if image_path:
                    local_img = self._download_image(image_path)
                    if local_img:
                        try:
                            # Add image to cell
                            xl_image = XLImage(local_img)
                            xl_image.width = 80
                            xl_image.height = 80
                            ws.add_image(xl_image, f'A{idx}')
                            ws.row_dimensions[idx].height = 90  # Adjust row height for image
                        except Exception as e:
                            self.logger.debug(f"Could not embed image: {e}")
                            ws['A' + str(idx)] = "[Image Error]"

            # Set header row height
            ws.row_dimensions[1].height = 25

            # Create output path
            if output_path is None:
                output_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    f"exports",
                    f"Articles_{self.timestamp}.xlsx"
                )

            # Create exports directory if needed
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Articles exported to Excel with images: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            raise

    def export_users_to_excel(self, users: List[Dict], output_path: str = None) -> str:
        """
        Export users to Excel file (without passwords)
        
        Args:
            users: List of user dictionaries
            output_path: Custom output path (optional)
        
        Returns:
            Path to created Excel file
        """
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl not installed")
            raise ImportError("openpyxl is required for Excel export")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"

            # Define styles
            header_fill = PatternFill(start_color=self.PRIMARY_HEX_EXCEL, end_color=self.PRIMARY_HEX_EXCEL, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ["ID", "Username", "Role", "Last Login", "Created Date"]
            ws.append(headers)

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Add data rows
            for user in users:
                row = [
                    user.get('id', '')[:8],
                    user.get('username', ''),
                    user.get('role', '').upper(),
                    user.get('last_login', 'Never')[:10] if user.get('last_login') else "Never",
                    user.get('created_at', '')[:10]
                ]
                ws.append(row)

                # Format data cells
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.row_dimensions[1].height = 25

            if output_path is None:
                output_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    f"exports",
                    f"Users_{self.timestamp}.xlsx"
                )

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            self.logger.info(f"Users exported to Excel: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            raise

    def export_articles_to_pdf(self, articles: List[Dict], output_path: str = None) -> str:
        """
        Export articles to PDF file with embedded images
        
        Args:
            articles: List of article dictionaries
            output_path: Custom output path (optional)
        
        Returns:
            Path to created PDF file
        """
        if not HAS_REPORTLAB:
            self.logger.error("reportlab not installed")
            raise ImportError("reportlab is required for PDF export")

        try:
            if output_path is None:
                output_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    f"exports",
                    f"Articles_{self.timestamp}.pdf"
                )

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Create PDF
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )

            # Styles with proper color
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.Color(*[c/255.0 for c in self.PRIMARY_RGB]),  # Convert RGB to 0-1 range
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_RIGHT,
                spaceAfter=20
            )

            # Build content
            story = []

            # Title
            story.append(Paragraph("NEXUZY ARTICAL - Articles Export", title_style))
            story.append(Paragraph(
                f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
                date_style
            ))
            story.append(Spacer(1, 0.3*inch))

            # Process articles with images
            for article in articles:
                # Article header
                story.append(Paragraph(
                    f"<b>{article.get('article_name', 'N/A')}</b> (ID: {article.get('id', '')[:8]})",
                    ParagraphStyle('ArticleTitle', parent=styles['Heading3'], fontSize=12, spaceAfter=10)
                ))

                # Try to embed image
                image_path = article.get('image_path')
                if image_path:
                    local_img = self._download_image(image_path, max_width=200, max_height=200)
                    if local_img:
                        try:
                            # Validate image before adding
                            img = Image.open(local_img)
                            img.verify()  # Check if valid image
                            
                            # Add image
                            rl_image = RLImage(local_img, width=2*inch, height=2*inch)
                            story.append(rl_image)
                            story.append(Spacer(1, 0.1*inch))
                        except Exception as e:
                            self.logger.debug(f"Could not add image to PDF: {e}")
                            story.append(Paragraph("[Image not available]", styles['Normal']))
                
                # Article details
                details = f"""
                <b>Mould:</b> {article.get('mould', 'N/A')}<br/>
                <b>Size:</b> {article.get('size', 'N/A')}<br/>
                <b>Gender:</b> {article.get('gender', 'N/A')}<br/>
                <b>Created:</b> {article.get('created_at', 'N/A')[:10]}<br/>
                <b>Status:</b> {'Synced' if article.get('sync_status', 0) == 1 else 'Pending'}
                """
                story.append(Paragraph(details, styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph("<hr/>", styles['Normal']))

            # Footer
            story.append(Spacer(1, 0.3*inch))
            footer_text = Paragraph(
                f"<i>Total Articles: {len(articles)} | Export generated by NEXUZY ARTICAL</i>",
                ParagraphStyle(
                    'Footer',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.grey,
                    alignment=TA_CENTER
                )
            )
            story.append(footer_text)

            # Build PDF
            doc.build(story)
            self.logger.info(f"Articles exported to PDF with images: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"PDF export failed: {e}")
            raise

    def export_users_to_pdf(self, users: List[Dict], output_path: str = None) -> str:
        """
        Export users to PDF file
        
        Args:
            users: List of user dictionaries
            output_path: Custom output path (optional)
        
        Returns:
            Path to created PDF file
        """
        if not HAS_REPORTLAB:
            self.logger.error("reportlab not installed")
            raise ImportError("reportlab is required for PDF export")

        try:
            if output_path is None:
                output_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    f"exports",
                    f"Users_{self.timestamp}.pdf"
                )

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.Color(*[c/255.0 for c in self.PRIMARY_RGB]),  # Convert RGB to 0-1 range
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            story = []
            story.append(Paragraph("NEXUZY ARTICAL - Users Report", title_style))
            story.append(Paragraph(
                f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
                ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_RIGHT, spaceAfter=20)
            ))
            story.append(Spacer(1, 0.3*inch))

            # Table data
            table_data = [[
                "Username", "Role", "Last Login", "Created Date", "Status"
            ]]

            for user in users:
                table_data.append([
                    user.get('username', ''),
                    user.get('role', '').upper(),
                    user.get('last_login', 'Never')[:10] if user.get('last_login') else "Never",
                    user.get('created_at', '')[:10],
                    "Active"
                ])

            table = Table(
                table_data,
                colWidths=[1.5*inch, 1*inch, 1.5*inch, 1.2*inch, 0.8*inch]
            )

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(*[c/255.0 for c in self.PRIMARY_RGB])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(self.SECONDARY_HEX_EXCEL)]),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))

            story.append(table)
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph(
                f"<i>Total Users: {len(users)} | Export generated by NEXUZY ARTICAL</i>",
                ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
            ))

            doc.build(story)
            self.logger.info(f"Users exported to PDF: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"PDF export failed: {e}")
            raise
